#File Handling in Python
"""
File Handling in Python: Concepts, Methods, Examples, Mini-project, and Interview Q&A

Definition:
-----------
File handling refers to how Python interacts with files stored on disk—reading data from files, writing data to files, and modifying file content.

Methods Commonly Used:
----------------------
- open()         : To open a file (returns a file object)
- read(), readline(), readlines()    : For reading file contents
- write(), writelines()              : For writing data to the file
- close()        : To close the file and free resources
- with statement (context manager): Safely open/close files
- seek(), tell(): Move/read file pointer position
- os and pathlib modules: For extra file operations (delete, rename, check existence, etc.)

File Modes:
-----------
- "r" : Read (default) - file must exist
- "w" : Write (truncates if file exists, else creates)
- "a" : Append (writes at end, creates if doesn't exist)
- "b" : Binary mode (add after above, eg "rb", "wb")
- "x" : Exclusive creation, fails if file exists
- "t" : Text mode (default, can combine with others)
- "+" : Update (read/write)

Syntax:
-------
# Open file for reading
f = open('myfile.txt', 'r')
data = f.read()
f.close()

# Using with statement (recommended)
with open('myfile.txt', 'r') as f:
    data = f.read()

# Open for writing (will overwrite)
with open('output.txt', 'w') as f:
    f.write('Hello, File!')

"""

# -----------------------------------------------
# Beginner Level Examples
# -----------------------------------------------

# Writing and reading a simple text file
with open('sample.txt', 'w') as f:
    f.write("Hello World!\nThis is a test file.\nPython can handle files easily.")

# Practice code: Using a local file path with Python file handling

# You can specify the full local path, or a relative path:
local_path = r"C:\Users\redrsa01\Documents\mydata.txt"  # Windows absolute path
# local_path = "/home/yourusername/mydata.txt"            # Linux/Mac absolute path
# local_path = "myfolder/myfile.txt"                      # Relative path


# Writing to the file using a local path
with open(local_path, 'w') as file:
    file.write("New Data\n"+"This is saved to my local path.\n")
    file.write("File path used: {}\n".format(local_path))
    file.write("     I am new learner for file handling concept in Python      ")



# After updating above, open the file and show the lines
# Read and print content of the file after writing

with open(local_path, 'r') as file:
    print("\n---Content of mydata.txt after writing---")
    data = file.read()
    print(data)


# Reading from the file at the same path
with open(local_path, 'r') as file:
    print("---Reading from local path---")
    for line in file:
        #print(line)
        print(line.strip())



# TIP: Use raw strings (prefix with r) for Windows paths to avoid escape issues
windows_path = r"C:\Users\redrsa01\Documents\mydata.txt"


with open(windows_path, 'w') as f:
    f.write("Raw string for Windows file path.")


# Example: using pathlib (modern and robust way for cross-platform paths)
from pathlib import Path


# Path object automatically handles separators correctly
myfile = r"C:\Users\redrsa01\Documents\python_file_handling_demo.txt"


with open(myfile, 'w') as f:
    f.write("Hello from pathlib! This file is created at:\n")
    f.write(str(myfile) + "\n")

with open(myfile, 'r') as f:
    print("\n---Content from pathlib path---")
    print(f.read())



# Reading the file back
with open(myfile, 'r') as f:
    content = f.read()
    print("---Content of sample.txt---")
    print(content,1)


# Reading line by line
with open(myfile, 'r') as f:
    for line in f:
        print("Line:", line.strip(),2)


# Appending to an existing file
with open(myfile, 'a') as f:
    f.write("Appended line at the end.")

with open(myfile, 'r') as f:
    print("\nAfter appending:")
    print(f.read())


# Handling Excel Files and Accessing Row/Column Data in Python

# One of the most popular libraries for Excel file manipulation is 'openpyxl' (for .xlsx files).
# You can install it via pip:
# pip install openpyxl
# To install 'openpyxl', you need to use the pip package installer.
# Run the following command in your terminal or command prompt (not in the Python shell):

# pip install openpyxl

# This will download and install 'openpyxl' and its dependencies.

# If you're using a Jupyter notebook, you can use:
# !pip install openpyxl

# Just make sure to install it before running code that imports or uses 'openpyxl'.

import openpyxl

# Create a new workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DataSheet"

# Writing headers and some sample data
ws.append(['Name', 'Age', 'Marks'])   # Adds a row
ws.append(['Alice', 21, 85])
ws.append(['Bob', 20, 92])
ws.append(['Charlie', 22, 78])

# Save the Excel file
wb.save('sample_data.xlsx')
print("Excel file 'sample_data.xlsx' created with sample data.")

# Now, let's load the file and play with row and column data
wb2 = openpyxl.load_workbook('sample_data.xlsx')
ws2 = wb2.active

# Accessing rows
print("\nReading all rows (as lists):")
for row in ws2.iter_rows(values_only=True):
    print(row)

# Accessing columns
print("\nReading all values from 'Marks' column:")
marks_col = ws2['C']  # Column C (third column)
for cell in marks_col[1:]:  # Skip header
    print(cell.value)

# Accessing specific cell
print("\nName in 2nd data row:", ws2.cell(row=3, column=1).value)  # (Row 3, Col 1): Bob

# Modifying a cell value
ws2.cell(row=2, column=3).value = 90  # Change Alice's mark to 90
wb2.save('sample_data_updated.xlsx')
print("\nUpdated Alice's marks and saved as 'sample_data_updated.xlsx'.")

# Summary: openpyxl allows you to read/write cell data, iterate over rows/columns, filter data, and much more with Excel files.



# -----------------------------------------------
# Intermediate Level Examples
# -----------------------------------------------

# Reading all lines into a list
with open(windows_path, 'r') as f:
    lines = f.readlines()
    line = f.readline()
print("1. Lines as list:", lines)
print("Type of lines:", type(lines))
print("2. Line as list:", line)
print("Type of line:", type(line))

# Writing multiple lines at once
lines_to_write = ["First line\n", "Second line\n", "Third line\n"]
with open('lines.txt', 'w') as f:
    f.writelines(lines_to_write)


# Reading one line at a time
with open('lines.txt', 'r') as f:
    line1 = f.readline()
    line2 = f.readline()
    print("First Line:", line1.strip())
    print("Second Line:", line2.strip())


# Getting file pointer location and seeking
with open('lines.txt', 'r') as f:
    print("Pointer at:", f.tell())
    print("Reading:", f.read(4))
    print("Pointer now:", f.tell())
    f.seek(0)
    print("After seek to start:", f.readline())



# Handling files that may not exist
try:
    with open('doesnotexist.txt', 'r') as f:
        data = f.read()
except FileNotFoundError:
    print("File doesnotexist.txt not found!")

# -----------------------------------------------
# Advanced Level Examples
# -----------------------------------------------

# Working with binary files: writing and reading
data_bytes = bytes([120, 3, 255, 0, 100])
with open('binaryfile.bin', 'wb') as f:
    f.write(data_bytes)


with open('binaryfile.bin', 'rb') as f:
    byte_content = f.read()
    print("Binary file content as bytes:", byte_content)

print("--------------------------------")
# Using 'with' for automatic file closure and error handling
def count_words(filename):
    try:
        with open(filename, 'r') as f:
            text = f.read()
            print("Text:", text)
            words = text.split()
            return len(words)
    except FileNotFoundError:
        print(f"{filename} not found!")
        return 0

#print("Number of words in 'sample.txt':", count_words(local_path))

# Using os module for file operations
import os

# Checking if a file exists
if os.path.exists('sample.txt'):
    print("'sample.txt' exists")

# Renaming a file
os.rename('lines.txt', 'renamed_lines.txt')
print("File renamed to renamed_lines.txt")

# Deleting a file
os.remove('binaryfile.bin')
print("binaryfile.bin deleted")

# -----------------------------------------------
# Mini Project: Simple Student Marks Saver and Reader
# -----------------------------------------------

# Program to save and read student marks using files

def save_marks(filename, records):
    """records: list of tuples (name, mark)"""
    with open(filename, 'w') as f:
        for name, mark in records:
            f.write(f"{name},{mark}\n")

def load_marks(filename):
    students = []
    with open(filename, 'r') as f:
        for line in f:
            name, mark = line.strip().split(',')
            students.append((name, int(mark)))
    return students

# Example usage:
student_records = [("Sai", 90), ("Amy", 76), ("Max", 85)]
save_marks('students.csv', student_records)

loaded = load_marks('students.csv')
print("\nLoaded student records from file:")
for name, mark in loaded:
    print(f"{name:10s} : {mark}")

# Find and print the student with the highest marks
topper = max(loaded, key=lambda x: x[1])
print(f"\nTopper: {topper[0]} with {topper[1]} marks")



'''
------------------------------------------------------------------
INTERVIEW QUESTIONS & ANSWERS (PYTHON FILE HANDLING)
------------------------------------------------------------------

Q1: What happens if you open a file in 'w' mode that already exists? 
A1: The file is truncated (emptied) before writing; previous content is erased.

Example:
with open('file.txt', 'w') as f:
    f.write('new data')
# file.txt now only contains 'new data'

------------------------------------------

Q2: How can you read a file's content safely ensuring the file is always closed?
A2: Use the 'with' statement (context manager); it ensures closure even if exceptions are raised.

Example:
with open('file.txt', 'r') as f:
    content = f.read()

------------------------------------------

Q3: How will you handle reading a huge file (GBs) without loading everything in memory?
A3: Process the file line by line using a for loop, avoiding .read() or .readlines():

with open('hugefile.txt', 'r') as f:
    for line in f:
        process(line)

------------------------------------------

Q4: What happens if you call write() in 'r' (read) mode?
A4: Raises io.UnsupportedOperation: not writable, since 'r' mode doesn't allow writing.

------------------------------------------

Q5: Distinguish between text(n) and binary(b) modes with an example.
A5: Text mode ('t', default) reads/writes str, handling encoding/decoding, \n translation. Binary mode ('b') reads/writes bytes as-is.

# Text mode:
with open('text.txt', 'w') as f:
    f.write('hello\n')

# Binary mode:
with open('bin.dat', 'wb') as f:
    f.write(b'hello\n')  # note leading b

------------------------------------------

Q6: How can you append to a file without losing its content?
A6: Use 'a' (append) mode:

with open('file.txt', 'a') as f:
    f.write('additional line\n')

------------------------------------------

Q7: Explain seek() and tell() with an example.
A7: tell() gives the current file pointer position (bytes). seek(offset, whence) moves the pointer to a desired position.

with open('file.txt') as f:
    f.read(10)
    print(f.tell())   # e.g., 10
    f.seek(0)
    print(f.tell())   # 0; moved to start

------------------------------------------

Q8: What's the best way to handle file-related errors (like missing files) in Python?
A8: Use try/except blocks, especially catching FileNotFoundError or IOError.

try:
    with open('nofile.txt') as f:
        f.read()
except FileNotFoundError:
    print("File not found!")

------------------------------------------

Q9: Demonstrate writing and reading structured data (e.g., CSV) using files.
A9:
# Writing CSV
rows = [("a", 1), ("b", 2)]
with open('data.csv', 'w') as f:
    for name, number in rows:
        f.write(f"{name},{number}\n")

# Reading CSV
with open('data.csv') as f:
    data = [tuple(line.strip().split(',')) for line in f]
    print(data)

------------------------------------------

Q10: Can you give a real-life scenario where you might use file handling in Python in automation?
A10: Logging application events to a file, saving user data or configuration, processing large logs for analytics, exporting/importing database dumps, etc.

Example:
# Saving logs
with open('log.txt', 'a') as f:
    f.write(f"{time.ctime()}: event X happened\n")

------------------------------------------------------------------
'''
